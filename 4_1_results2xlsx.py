import os
import openpyxl

# model_names = ["Ours", "MLP", "LSTM", "TCN", "CNN", "iTransformer", "TimesNet", "DLinear", "Autoformer",
#                "PatchTST", "SparseTSF", "Informer", "TimeMixer"]
model_names = ["Ours"]
capacities = [15.5, 20, 23, 30, 52, 63, 67, 105]
# capacities = [15.5, 20]

flag = 'benchmark'
# flag = 'zero_shot'

if flag == 'benchmark':
    result_path = './outputs/{model_name}_benchmark_{capacity}.txt'
    save_path = "./results_benchmark.xlsx"
elif flag == 'zero_shot':
    result_path = './outputs/{model_name}_zero_shot_{capacity}.txt'
    save_path = "./results_zero_shot.xlsx"

# ========== 创建工作簿 ==========
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Results"

# 第一行：容量
ws.cell(1, 1, "")  # 左上角空
col = 2
for cap in capacities:
    ws.cell(1, col, cap)
    col += 3  # 占3列（RMSE, MAE, MAPE）

# 第二行：指标
col = 2
for _ in capacities:
    ws.cell(2, col, "RMSE")
    ws.cell(2, col + 1, "MAE")
    ws.cell(2, col + 2, "MAPE")
    col += 3

# ========== 读取结果 ==========
row = 3
for model in model_names:
    ws.cell(row, 1, model)  # 第一列写模型名
    col = 2
    for cap in capacities:
        file_path = result_path.format(model_name=model, capacity=cap)
        if os.path.exists(file_path):
            with open(file_path, "r") as f:
                lines = f.readlines()
                if len(lines) >= 2:
                    values_line = lines[-1].strip()
                    parts = values_line.split()
                    if len(parts) == 3:
                        rmse, mae, mape = parts
                        ws.cell(row, col, float(rmse))
                        ws.cell(row, col + 1, float(mae))
                        # MAPE 保留百分号（字符串形式）
                        if not mape.endswith("%"):
                            mape = mape + "%"
                        ws.cell(row, col + 2, mape)
        col += 3
    row += 1

# 保存文件
wb.save(save_path)
print(f"结果已保存到 {save_path}")
